"""
Extrai valores de deposito recursal dos PDFs - lista TRABALHADOR RECORREU (Uber parte re).
Gera Excel consolidado final com AMBAS as listas.

Uso:
  python3 scripts/captura/acervo-geral/extrair-valores-pdfs-parte-re.py
"""

import json
import os
import re
from collections import defaultdict
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Diretórios
RESULTS_DIR_RE = os.path.join(os.path.dirname(__file__), "..", "..", "results", "depositos-uber-tst-parte-re")
RESULTS_DIR_AUTORA = os.path.join(os.path.dirname(__file__), "..", "..", "results", "depositos-uber-tst")
PDFS_DIR = os.path.join(RESULTS_DIR_RE, "pdfs")
MAPEAMENTO_PATH = os.path.join(RESULTS_DIR_RE, "mapeamento-docs.json")
OUTPUT_EXCEL = os.path.expanduser("~/Desktop/QUADRO_COMPLETO_depositos_uber_tst.xlsx")

VALOR_PATTERNS = [
    r'[Vv]alor[\s:]*R\$\s*([\d.,]+)',
    r'R\$\s*([\d]+(?:\.[\d]{3})*(?:,[\d]{2}))',
    r'(?:valor|depósito|deposito|total)[\s:]*(?:R\$\s*)?([\d]+(?:\.[\d]{3})*,[\d]{2})',
    r'[Vv]alor\s+(?:do\s+)?[Dd]ep[oó]sito[\s:]*R?\$?\s*([\d.,]+)',
]


def extrair_valor_pdf(pdf_path):
    resultado = {"valores_encontrados": [], "valor_principal": None, "erro": None}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto = ""
            for page in pdf.pages:
                texto += (page.extract_text() or "") + "\n"
            todos_valores = []
            for pattern in VALOR_PATTERNS:
                for match in re.findall(pattern, texto, re.IGNORECASE):
                    try:
                        v = float(match.strip().replace(".", "").replace(",", "."))
                        if 0 < v < 1_000_000:
                            todos_valores.append(v)
                    except ValueError:
                        continue
            valores_unicos = sorted(set(todos_valores), reverse=True)
            resultado["valores_encontrados"] = valores_unicos
            if valores_unicos:
                resultado["valor_principal"] = valores_unicos[0]
    except Exception as e:
        resultado["erro"] = str(e)
    return resultado


def processar_lista(mapeamento_path, pdfs_dir, label):
    """Processa uma lista de PDFs e retorna dados agrupados por processo."""
    print(f"\n{'='*60}")
    print(f"  {label}")
    print(f"{'='*60}")

    with open(mapeamento_path, "r") as f:
        mapeamento = json.load(f)

    docs_com_pdf = [d for d in mapeamento if d.get("pdf_path")]
    print(f"  Docs: {len(mapeamento)} | Com PDF: {len(docs_com_pdf)}")

    extraidos_ok = 0
    for i, doc in enumerate(docs_com_pdf):
        if i % 100 == 0 or i == len(docs_com_pdf) - 1:
            print(f"  Progresso: {i + 1}/{len(docs_com_pdf)}")
        pdf_path = os.path.join(pdfs_dir, doc["pdf_path"])
        if not os.path.exists(pdf_path):
            doc["valor_extraido"] = None
            continue
        extracao = extrair_valor_pdf(pdf_path)
        doc["valor_extraido"] = extracao["valor_principal"]
        if extracao["valor_principal"]:
            extraidos_ok += 1

    print(f"  Valores extraidos: {extraidos_ok}/{len(docs_com_pdf)}")

    # Agrupar por processo
    por_processo = defaultdict(lambda: {
        "contraparte": "", "classe_judicial": "",
        "dep_recursal_1g": 0.0, "dep_recursal_2g": 0.0,
        "custas_1g": 0.0, "custas_2g": 0.0,
    })

    for doc in docs_com_pdf:
        nproc = doc["numero_processo"]
        p = por_processo[nproc]
        p["contraparte"] = doc.get("nome_parte_autora") or doc.get("nome_parte_re", "")
        p["classe_judicial"] = doc.get("classe_judicial", "")

        inst = doc.get("instancia", "")
        valor = doc.get("valor_extraido") or 0
        if valor <= 0:
            continue

        titulo = (doc.get("titulo") or "").lower()
        tipo = (doc.get("tipo") or "").lower()
        is_custas = "gru" in tipo or "recolhimento" in tipo or ("custas" in titulo and "depósito" not in titulo)
        is_1g = "1" in inst
        is_2g = "2" in inst

        if is_custas:
            if is_1g: p["custas_1g"] = max(p["custas_1g"], valor)
            elif is_2g: p["custas_2g"] = max(p["custas_2g"], valor)
        else:
            if is_1g: p["dep_recursal_1g"] = max(p["dep_recursal_1g"], valor)
            elif is_2g: p["dep_recursal_2g"] = max(p["dep_recursal_2g"], valor)

    return dict(por_processo)


def main():
    print("=" * 80)
    print("QUADRO COMPLETO — DEPOSITOS RECURSAIS UBER NO TST")
    print("=" * 80)

    # Lista 1: Uber recorreu (parte autora no TST)
    dados_uber_recorreu = processar_lista(
        os.path.join(RESULTS_DIR_AUTORA, "mapeamento-docs.json"),
        os.path.join(RESULTS_DIR_AUTORA, "pdfs"),
        "LISTA 1: UBER RECORREU (parte autora no TST)"
    )

    # Lista 2: Trabalhador recorreu (Uber parte ré no TST)
    dados_trab_recorreu = processar_lista(
        MAPEAMENTO_PATH, PDFS_DIR,
        "LISTA 2: TRABALHADOR RECORREU (Uber parte re no TST)"
    )

    # ─── EXCEL CONSOLIDADO ───
    print(f"\nGerando Excel consolidado: {OUTPUT_EXCEL}")

    wb = Workbook()
    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    df = Font(name="Calibri", size=10)
    da = Alignment(vertical="center", wrap_text=True)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    money_fmt = '#,##0.00'
    bf = Font(name="Calibri", bold=True, size=11)

    def write_sheet(ws, dados, header_fill, alt_fill, contraparte_label):
        headers = [
            "No", "Numero do Processo", "Classe", contraparte_label,
            "Dep. Recursal 1G (R$)", "Custas 1G (R$)",
            "Dep. Recursal 2G (R$)", "Custas 2G (R$)",
            "Total Depositado (R$)"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hf
            cell.fill = header_fill
            cell.alignment = ha
            cell.border = tb

        processos = sorted(dados.items(),
            key=lambda x: x[1]["dep_recursal_1g"] + x[1]["dep_recursal_2g"] + x[1]["custas_1g"] + x[1]["custas_2g"],
            reverse=True)

        for i, (nproc, p) in enumerate(processos, 1):
            total = p["dep_recursal_1g"] + p["dep_recursal_2g"] + p["custas_1g"] + p["custas_2g"]
            values = [
                i, nproc, p["classe_judicial"], p["contraparte"],
                p["dep_recursal_1g"] if p["dep_recursal_1g"] > 0 else "",
                p["custas_1g"] if p["custas_1g"] > 0 else "",
                p["dep_recursal_2g"] if p["dep_recursal_2g"] > 0 else "",
                p["custas_2g"] if p["custas_2g"] > 0 else "",
                total if total > 0 else "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=i+1, column=col, value=val)
                cell.font = df
                cell.alignment = da
                cell.border = tb
                if i % 2 == 0: cell.fill = alt_fill
                if col in (5,6,7,8,9) and isinstance(val, float) and val > 0:
                    cell.number_format = money_fmt

        # Linha total
        r = len(processos) + 2
        ws.cell(row=r, column=2, value="TOTAL").font = bf
        for col in range(1, len(headers)+1):
            ws.cell(row=r, column=col).border = tb
            ws.cell(row=r, column=col).font = bf
        ws.cell(row=r, column=5, value=sum(p["dep_recursal_1g"] for _,p in processos)).number_format = money_fmt
        ws.cell(row=r, column=6, value=sum(p["custas_1g"] for _,p in processos)).number_format = money_fmt
        ws.cell(row=r, column=7, value=sum(p["dep_recursal_2g"] for _,p in processos)).number_format = money_fmt
        ws.cell(row=r, column=8, value=sum(p["custas_2g"] for _,p in processos)).number_format = money_fmt
        total_geral = sum(p["dep_recursal_1g"]+p["dep_recursal_2g"]+p["custas_1g"]+p["custas_2g"] for _,p in processos)
        ws.cell(row=r, column=9, value=total_geral).number_format = money_fmt

        widths = [5, 30, 8, 35, 18, 15, 18, 15, 20]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+col)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{len(processos)+1}"

        return total_geral, processos

    # Aba 1: Uber recorreu
    ws1 = wb.active
    ws1.title = "Uber Recorreu (411)"
    fill1 = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    alt1 = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    total1, procs1 = write_sheet(ws1, dados_uber_recorreu, fill1, alt1, "Reclamante")

    # Aba 2: Trabalhador recorreu
    ws2 = wb.create_sheet("Trab. Recorreu (478)")
    fill2 = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt2 = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    total2, procs2 = write_sheet(ws2, dados_trab_recorreu, fill2, alt2, "Reclamante")

    # Aba 3: Resumo consolidado
    ws3 = wb.create_sheet("Resumo Consolidado")
    ws3.column_dimensions["A"].width = 55
    ws3.column_dimensions["B"].width = 25

    # Totais por categoria
    dep1_uber = sum(p["dep_recursal_1g"] for p in dados_uber_recorreu.values())
    dep2_uber = sum(p["dep_recursal_2g"] for p in dados_uber_recorreu.values())
    cst1_uber = sum(p["custas_1g"] for p in dados_uber_recorreu.values())
    cst2_uber = sum(p["custas_2g"] for p in dados_uber_recorreu.values())

    dep1_trab = sum(p["dep_recursal_1g"] for p in dados_trab_recorreu.values())
    dep2_trab = sum(p["dep_recursal_2g"] for p in dados_trab_recorreu.values())
    cst1_trab = sum(p["custas_1g"] for p in dados_trab_recorreu.values())
    cst2_trab = sum(p["custas_2g"] for p in dados_trab_recorreu.values())

    total_dep_recursal = dep1_uber + dep2_uber + dep1_trab + dep2_trab
    total_custas = cst1_uber + cst2_uber + cst1_trab + cst2_trab
    total_tudo = total_dep_recursal + total_custas

    procs_valor_1 = sum(1 for p in dados_uber_recorreu.values() if p["dep_recursal_1g"]+p["dep_recursal_2g"]+p["custas_1g"]+p["custas_2g"] > 0)
    procs_valor_2 = sum(1 for p in dados_trab_recorreu.values() if p["dep_recursal_1g"]+p["dep_recursal_2g"]+p["custas_1g"]+p["custas_2g"] > 0)

    resumo = [
        ("QUADRO COMPLETO — DEPOSITOS RECURSAIS UBER NO TST", ""),
        ("Data da analise", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Fonte", "Raspagem direta PJE TST + extracao PDF"),
        ("", ""),
        ("═══ LISTA 1: UBER RECORREU AO TST (parte autora) ═══", ""),
        ("Total processos", "411"),
        ("Processos com docs de deposito", str(len(dados_uber_recorreu))),
        ("Processos com valor extraido", str(procs_valor_1)),
        ("Dep. recursal 1a instancia", f"R$ {dep1_uber:,.2f}"),
        ("Dep. recursal 2a instancia", f"R$ {dep2_uber:,.2f}"),
        ("Custas 1a instancia", f"R$ {cst1_uber:,.2f}"),
        ("Custas 2a instancia", f"R$ {cst2_uber:,.2f}"),
        ("Subtotal Uber recorreu", f"R$ {total1:,.2f}"),
        ("", ""),
        ("═══ LISTA 2: TRABALHADOR RECORREU AO TST (Uber parte re) ═══", ""),
        ("Total processos", "478"),
        ("Processos com docs de deposito", str(len(dados_trab_recorreu))),
        ("Processos com valor extraido", str(procs_valor_2)),
        ("Dep. recursal 1a instancia", f"R$ {dep1_trab:,.2f}"),
        ("Dep. recursal 2a instancia", f"R$ {dep2_trab:,.2f}"),
        ("Custas 1a instancia", f"R$ {cst1_trab:,.2f}"),
        ("Custas 2a instancia", f"R$ {cst2_trab:,.2f}"),
        ("Subtotal Trabalhador recorreu", f"R$ {total2:,.2f}"),
        ("", ""),
        ("═══ TOTAIS CONSOLIDADOS ═══", ""),
        ("TOTAL DEPOSITOS RECURSAIS", f"R$ {total_dep_recursal:,.2f}"),
        ("TOTAL CUSTAS", f"R$ {total_custas:,.2f}"),
        ("", ""),
        ("TOTAL GERAL DEPOSITADO PELA UBER", f"R$ {total_tudo:,.2f}"),
    ]

    tf = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    sf = Font(name="Calibri", bold=True, size=11, color="8B0000")
    hl = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    hl2 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for row_idx, (a, b) in enumerate(resumo, 1):
        ws3.cell(row=row_idx, column=1, value=a)
        ws3.cell(row=row_idx, column=2, value=b)
        if row_idx == 1:
            ws3.cell(row=row_idx, column=1).font = tf
        elif a.startswith("═══"):
            ws3.cell(row=row_idx, column=1).font = sf
        elif "TOTAL GERAL" in a:
            ws3.cell(row=row_idx, column=1).font = Font(name="Calibri", bold=True, size=13, color="8B0000")
            ws3.cell(row=row_idx, column=2).font = Font(name="Calibri", bold=True, size=13, color="8B0000")
            ws3.cell(row=row_idx, column=1).fill = hl2
            ws3.cell(row=row_idx, column=2).fill = hl2
        elif a.startswith("Subtotal") or a.startswith("TOTAL "):
            ws3.cell(row=row_idx, column=1).font = bf
            ws3.cell(row=row_idx, column=2).font = bf
            ws3.cell(row=row_idx, column=1).fill = hl
            ws3.cell(row=row_idx, column=2).fill = hl

    wb.save(OUTPUT_EXCEL)

    print(f"\n{'='*80}")
    print("QUADRO COMPLETO")
    print(f"{'='*80}")
    print(f"  Uber recorreu:           R$ {total1:,.2f} ({procs_valor_1} processos)")
    print(f"  Trabalhador recorreu:    R$ {total2:,.2f} ({procs_valor_2} processos)")
    print(f"  TOTAL DEP. RECURSAIS:    R$ {total_dep_recursal:,.2f}")
    print(f"  TOTAL CUSTAS:            R$ {total_custas:,.2f}")
    print(f"  TOTAL GERAL:             R$ {total_tudo:,.2f}")
    print(f"  Excel:                   {OUTPUT_EXCEL}")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    main()
