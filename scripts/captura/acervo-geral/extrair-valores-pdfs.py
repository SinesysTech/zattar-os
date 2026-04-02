"""
Script para extrair valores de deposito recursal dos PDFs baixados.
Le os PDFs da pasta pdfs/, extrai valores com regex, e gera Excel final.

Uso:
  python3 scripts/captura/acervo-geral/extrair-valores-pdfs.py
"""

import json
import os
import re
from collections import defaultdict
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "results", "depositos-uber-tst")
PDFS_DIR = os.path.join(RESULTS_DIR, "pdfs")
MAPEAMENTO_PATH = os.path.join(RESULTS_DIR, "mapeamento-docs.json")
OUTPUT_EXCEL = os.path.expanduser("~/Desktop/processos_uber_tst_depositos_recursais_FINAL.xlsx")

# Regex para extrair valores monetarios de PDFs de deposito
VALOR_PATTERNS = [
    # Padrao SISCONDJ / Banco do Brasil: "Valor: R$ 13.133,46" ou "VALOR R$ 13.133,46"
    r'[Vv]alor[\s:]*R\$\s*([\d.,]+)',
    # "R$ 13.133,46" generico
    r'R\$\s*([\d]+(?:\.[\d]{3})*(?:,[\d]{2}))',
    # Valor sem R$: "13.133,46" precedido de "valor" ou "deposito"
    r'(?:valor|depósito|deposito|total)[\s:]*(?:R\$\s*)?([\d]+(?:\.[\d]{3})*,[\d]{2})',
    # "Valor do Deposito: 13133.46" (formato com ponto decimal)
    r'[Vv]alor\s+(?:do\s+)?[Dd]ep[oó]sito[\s:]*R?\$?\s*([\d.,]+)',
]


def extrair_valor_pdf(pdf_path: str) -> dict:
    """Extrai texto do PDF e busca valor de deposito."""
    resultado = {
        "texto_extraido": "",
        "valores_encontrados": [],
        "valor_principal": None,
        "erro": None,
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                texto = page.extract_text() or ""
                texto_completo += texto + "\n"

            resultado["texto_extraido"] = texto_completo[:2000]  # Limitar para nao explodir JSON

            # Buscar valores com cada pattern
            todos_valores = []
            for pattern in VALOR_PATTERNS:
                matches = re.findall(pattern, texto_completo, re.IGNORECASE)
                for match in matches:
                    # Normalizar valor
                    valor_str = match.strip()
                    try:
                        # Formato brasileiro: 13.133,46 -> 13133.46
                        valor_normalizado = valor_str.replace(".", "").replace(",", ".")
                        valor = float(valor_normalizado)
                        if valor > 0 and valor < 1_000_000:  # Sanidade
                            todos_valores.append(valor)
                    except ValueError:
                        continue

            # Deduplicar e ordenar (maior primeiro)
            valores_unicos = sorted(set(todos_valores), reverse=True)
            resultado["valores_encontrados"] = valores_unicos

            # O valor principal eh o MAIOR valor encontrado
            # (deposito recursal eh normalmente o maior; custas sao menores)
            if valores_unicos:
                resultado["valor_principal"] = valores_unicos[0]

    except Exception as e:
        resultado["erro"] = str(e)

    return resultado


def classificar_documento(tipo: str, titulo: str) -> str:
    """Classifica se eh deposito recursal ou custas."""
    titulo_lower = titulo.lower()
    tipo_lower = tipo.lower()

    if "gru" in tipo_lower or "recolhimento" in tipo_lower:
        return "custas"
    if "custas" in titulo_lower and "depósito" not in titulo_lower:
        return "custas"
    if "gru" in titulo_lower:
        return "custas"
    return "deposito_recursal"


def main():
    print("=" * 80)
    print("EXTRACAO DE VALORES DE DEPOSITOS RECURSAIS - UBER TST")
    print("=" * 80)

    # 1. Carregar mapeamento
    print(f"\n1. Carregando mapeamento: {MAPEAMENTO_PATH}")
    with open(MAPEAMENTO_PATH, "r") as f:
        mapeamento = json.load(f)

    docs_com_pdf = [d for d in mapeamento if d.get("pdf_path")]
    print(f"   Total: {len(mapeamento)} docs, {len(docs_com_pdf)} com PDF baixado")

    # 2. Extrair valores de cada PDF
    print(f"\n2. Extraindo valores de {len(docs_com_pdf)} PDFs...")

    resultados_extracao = []
    extraidos_ok = 0
    sem_valor = 0
    com_erro = 0

    for i, doc in enumerate(docs_com_pdf):
        if i % 100 == 0 or i == len(docs_com_pdf) - 1:
            print(f"   Progresso: {i + 1}/{len(docs_com_pdf)}")

        pdf_path = os.path.join(PDFS_DIR, doc["pdf_path"])
        if not os.path.exists(pdf_path):
            doc["extracao"] = {"erro": "PDF nao encontrado"}
            com_erro += 1
            continue

        extracao = extrair_valor_pdf(pdf_path)
        doc["extracao"] = extracao
        doc["valor_extraido"] = extracao["valor_principal"]
        doc["classificacao"] = classificar_documento(doc["tipo"], doc["titulo"])

        if extracao["valor_principal"]:
            extraidos_ok += 1
        elif extracao["erro"]:
            com_erro += 1
        else:
            sem_valor += 1

        resultados_extracao.append(doc)

    print(f"\n   Valores extraidos: {extraidos_ok}")
    print(f"   Sem valor encontrado: {sem_valor}")
    print(f"   Com erro: {com_erro}")

    # 3. Agrupar por processo
    print("\n3. Agrupando por processo...")

    por_processo = defaultdict(lambda: {
        "nome_parte_re": "",
        "classe_judicial": "",
        "dep_recursal_1g": 0.0,
        "dep_recursal_2g": 0.0,
        "custas_1g": 0.0,
        "custas_2g": 0.0,
        "docs_1g": 0,
        "docs_2g": 0,
        "docs_sem_valor": 0,
        "todos_valores": [],
        "detalhes": [],
    })

    for doc in resultados_extracao:
        nproc = doc["numero_processo"]
        p = por_processo[nproc]
        p["nome_parte_re"] = doc["nome_parte_re"]
        p["classe_judicial"] = doc["classe_judicial"]

        instancia = doc.get("instancia", "")
        is_1g = "1" in instancia
        is_2g = "2" in instancia
        valor = doc.get("valor_extraido") or 0
        classificacao = doc.get("classificacao", "deposito_recursal")

        detalhe = f"{doc['titulo']} ({instancia}) = R$ {valor:,.2f}" if valor else f"{doc['titulo']} ({instancia}) = SEM VALOR"
        p["detalhes"].append(detalhe)

        if valor <= 0:
            p["docs_sem_valor"] += 1
            continue

        p["todos_valores"].append({"valor": valor, "instancia": instancia, "tipo": classificacao})

        if classificacao == "custas":
            if is_1g:
                p["custas_1g"] = max(p["custas_1g"], valor)
                p["docs_1g"] += 1
            elif is_2g:
                p["custas_2g"] = max(p["custas_2g"], valor)
                p["docs_2g"] += 1
        else:
            if is_1g:
                p["dep_recursal_1g"] = max(p["dep_recursal_1g"], valor)
                p["docs_1g"] += 1
            elif is_2g:
                p["dep_recursal_2g"] = max(p["dep_recursal_2g"], valor)
                p["docs_2g"] += 1

    print(f"   {len(por_processo)} processos com documentos")

    # 4. Gerar Excel
    print(f"\n4. Gerando Excel: {OUTPUT_EXCEL}")

    wb = Workbook()

    # Styles
    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    df = Font(name="Calibri", size=10)
    da = Alignment(vertical="center", wrap_text=True)
    alt = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    money_fmt = '#,##0.00'

    headers = [
        "No", "Numero do Processo", "Classe", "Reclamante",
        "Uber Perdeu Em", "Dep. Recursal 1G (R$)", "Custas 1G (R$)",
        "Dep. Recursal 2G (R$)", "Custas 2G (R$)",
        "Total Depositado (R$)", "Docs Analisados", "Docs Sem Valor",
    ]

    ws = wb.active
    ws.title = "Depositos Recursais"

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        cell.border = tb

    # Ordenar por total depositado (maior primeiro)
    processos_ordenados = sorted(
        por_processo.items(),
        key=lambda x: x[1]["dep_recursal_1g"] + x[1]["dep_recursal_2g"] + x[1]["custas_1g"] + x[1]["custas_2g"],
        reverse=True,
    )

    total_geral = 0
    total_dep_recursal = 0
    total_custas = 0

    for i, (nproc, p) in enumerate(processos_ordenados, 1):
        dep_1g = p["dep_recursal_1g"]
        dep_2g = p["dep_recursal_2g"]
        cst_1g = p["custas_1g"]
        cst_2g = p["custas_2g"]
        total = dep_1g + dep_2g + cst_1g + cst_2g

        total_geral += total
        total_dep_recursal += dep_1g + dep_2g
        total_custas += cst_1g + cst_2g

        tem_1g = dep_1g > 0 or cst_1g > 0
        tem_2g = dep_2g > 0 or cst_2g > 0
        if tem_1g and tem_2g:
            perdeu = "1a e 2a instancia"
        elif tem_1g:
            perdeu = "1a instancia"
        elif tem_2g:
            perdeu = "2a instancia"
        else:
            perdeu = "Sem valor extraido"

        values = [
            i, nproc, p["classe_judicial"], p["nome_parte_re"],
            perdeu,
            dep_1g if dep_1g > 0 else "",
            cst_1g if cst_1g > 0 else "",
            dep_2g if dep_2g > 0 else "",
            cst_2g if cst_2g > 0 else "",
            total if total > 0 else "",
            p["docs_1g"] + p["docs_2g"],
            p["docs_sem_valor"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.font = df
            cell.alignment = da
            cell.border = tb
            if i % 2 == 0:
                cell.fill = alt
            if col in (6, 7, 8, 9, 10) and isinstance(val, float) and val > 0:
                cell.number_format = money_fmt

    # Linha de totais
    total_row = len(processos_ordenados) + 2
    ws.cell(row=total_row, column=1, value="").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).border = tb
        ws.cell(row=total_row, column=col).font = Font(name="Calibri", bold=True, size=11)

    sum_dep_1g = sum(p["dep_recursal_1g"] for _, p in processos_ordenados)
    sum_cst_1g = sum(p["custas_1g"] for _, p in processos_ordenados)
    sum_dep_2g = sum(p["dep_recursal_2g"] for _, p in processos_ordenados)
    sum_cst_2g = sum(p["custas_2g"] for _, p in processos_ordenados)

    ws.cell(row=total_row, column=6, value=sum_dep_1g).number_format = money_fmt
    ws.cell(row=total_row, column=7, value=sum_cst_1g).number_format = money_fmt
    ws.cell(row=total_row, column=8, value=sum_dep_2g).number_format = money_fmt
    ws.cell(row=total_row, column=9, value=sum_cst_2g).number_format = money_fmt
    ws.cell(row=total_row, column=10, value=total_geral).number_format = money_fmt

    # Larguras
    widths = [5, 30, 8, 35, 22, 18, 15, 18, 15, 20, 10, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else ""].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(processos_ordenados) + 1}"

    # Aba Resumo
    ws2 = wb.create_sheet("Resumo")
    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 25

    processos_com_valor = sum(1 for _, p in processos_ordenados if p["dep_recursal_1g"] + p["dep_recursal_2g"] + p["custas_1g"] + p["custas_2g"] > 0)
    processos_ambos = sum(1 for _, p in processos_ordenados if (p["dep_recursal_1g"] > 0 or p["custas_1g"] > 0) and (p["dep_recursal_2g"] > 0 or p["custas_2g"] > 0))
    processos_so_1g = sum(1 for _, p in processos_ordenados if (p["dep_recursal_1g"] > 0 or p["custas_1g"] > 0) and p["dep_recursal_2g"] == 0 and p["custas_2g"] == 0)
    processos_so_2g = sum(1 for _, p in processos_ordenados if p["dep_recursal_1g"] == 0 and p["custas_1g"] == 0 and (p["dep_recursal_2g"] > 0 or p["custas_2g"] > 0))

    resumo = [
        ("DEPOSITOS RECURSAIS - UBER NO TST", ""),
        ("Data da analise", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Total processos Uber recorreu", "411"),
        ("Processos com docs de deposito", str(len(por_processo))),
        ("Processos com valor extraido do PDF", str(processos_com_valor)),
        ("", ""),
        ("ONDE A UBER PERDEU:", ""),
        ("Perdeu em 1a e 2a instancia (2 depositos)", str(processos_ambos)),
        ("Perdeu so em 1a instancia", str(processos_so_1g)),
        ("Perdeu so em 2a instancia", str(processos_so_2g)),
        ("", ""),
        ("VALORES EXTRAIDOS DOS PDFs:", ""),
        ("Total dep. recursal 1a instancia", f"R$ {sum_dep_1g:,.2f}"),
        ("Total custas 1a instancia", f"R$ {sum_cst_1g:,.2f}"),
        ("Total dep. recursal 2a instancia", f"R$ {sum_dep_2g:,.2f}"),
        ("Total custas 2a instancia", f"R$ {sum_cst_2g:,.2f}"),
        ("", ""),
        ("TOTAL DEPOSITOS RECURSAIS", f"R$ {total_dep_recursal:,.2f}"),
        ("TOTAL CUSTAS", f"R$ {total_custas:,.2f}"),
        ("TOTAL GERAL DEPOSITADO", f"R$ {total_geral:,.2f}"),
    ]

    title_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    bold_font = Font(name="Calibri", bold=True, size=11)
    hl = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for row_idx, (a, b) in enumerate(resumo, 1):
        ws2.cell(row=row_idx, column=1, value=a)
        ws2.cell(row=row_idx, column=2, value=b)
        if row_idx == 1:
            ws2.cell(row=row_idx, column=1).font = title_font
        elif "TOTAL" in a and a.startswith("TOTAL"):
            ws2.cell(row=row_idx, column=1).font = bold_font
            ws2.cell(row=row_idx, column=2).font = bold_font
            ws2.cell(row=row_idx, column=1).fill = hl
            ws2.cell(row=row_idx, column=2).fill = hl

    wb.save(OUTPUT_EXCEL)

    # Salvar JSON com resultados completos
    json_path = os.path.join(RESULTS_DIR, "resultados-extracao.json")
    with open(json_path, "w") as f:
        # Remover texto extraido para nao explodir o JSON
        for doc in resultados_extracao:
            if "extracao" in doc:
                doc["extracao"].pop("texto_extraido", None)
        json.dump(resultados_extracao, f, ensure_ascii=False, indent=2)

    print(f"\n{'=' * 80}")
    print("RESUMO FINAL")
    print(f"{'=' * 80}")
    print(f"   Processos analisados:      {len(por_processo)}")
    print(f"   Com valor extraido:        {processos_com_valor}")
    print(f"   Total dep. recursal:       R$ {total_dep_recursal:,.2f}")
    print(f"   Total custas:              R$ {total_custas:,.2f}")
    print(f"   TOTAL GERAL:               R$ {total_geral:,.2f}")
    print(f"   Excel:                     {OUTPUT_EXCEL}")
    print(f"{'=' * 80}\n")


if __name__ == "__main__":
    main()
